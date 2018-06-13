"""
bt_expense.py
~~~~~~~~~
Pull expenses from Excel Spreadsheet and upload to BigTime via REST HTTP call.
"""
import os
import json
from pprint import pprint as pp
# from pprint import pformat as pf
import requests as r
from openpyxl import load_workbook
import subprocess
import win32com.client as winclient

# CD
# os.chdir('bt_expense')

# Constants
BASE = 'https://iq.bigtime.net/BigtimeData/api/v2'
UTF = 'utf-8'
# Global Variables
BT_LOOKUP = {'proj': {},
             'cat': {}, }


class Authorizer(object):
    """authorizes a BitTime REST API session.
    Can authorize using a user login and password or an API key.

    User login and password will be used to obtain an API key.
    If API key is provided, skip the step of obtaining API key"""

    def __init__(self, workbook_filename='Expenses.xlsx', staffsid=None):
        self.wb_name = workbook_filename
        self.staffsid = staffsid
        self.auth_header = self._build_credentials()
        self.userid = self.auth_header['userid']
        self.userpwd = self.auth_header['pwd']
        self.api_key = None
        self._authorized = False
        self.header = self.authorize_session()

    def _build_credentials(self):
        """Pulls Login information from the `Setup` worksheet. Return dictionary
        for Auth Header."""
        keys = get_values('Setup', 'A1', 'A4', workbook_name=self.wb_name)
        values = get_values('Setup', 'B1', 'B4', workbook_name=self.wb_name)
        header = {k: v for (k, v) in zip(keys, values)}
        header['Content-Type'] = 'application/json'
        return header

    def authorize_session(self):
        response = r.post('{}/session'.format(BASE),
                          headers={'Content-Type': 'application/json'},
                          data=json.dumps(self.auth_header).encode('utf-8'))
        if str(response.status_code)[0] is not '2':
            # TODO Raise Requests HTTP Error
            raise(ConnectionRefusedError)
        response_dict = json.loads(response.text)
        self.api_key = response_dict['token']
        self.staffsid = response_dict['staffsid']
        header = {'X-Auth-Token': self.api_key,
                  'X-Auth-Realm': self.auth_header['Firm'],
                  'Content-Type': self.auth_header['Content-Type']}
        self._authorized = True
        # print('Session Header\n', header)
        return header

class Expensor(Authorizer):

    def prep_expenses(self, save=True):
        pnames = get_values('Expenses', 'A2', 'A102')
        projs = get_values('Expenses', 'F2', 'F102')
        cats = get_values('Expenses', 'G2', 'G102')
        dates = get_values('Expenses', 'C2', 'C102')
        costs = get_values('Expenses', 'D2', 'D102')
        notes = get_values('Expenses', 'E2', 'E102')
        expense_entries = []
        total_cost = 0
        daily_food_costs = {}
        for proj, cat, date, cost, note, pname in zip(projs, cats,
                                                      dates, costs,
                                                      notes, pnames):
            
            #Get per diem value from the workbook
            per_diem = get_values('Setup', 'B6', 'B6')[0]

            if cost and date:
                if per_diem:
                    # perform per diem validation
                    total_cost, cost, daily_food_costs = self._per_diem_validation(daily_food_costs, cat,\
                    date, cost, note, per_diem, total_cost)
                content = {'staffsid': int(self.staffsid),
                           'projectsid': int(proj),
                           'catsid': int(cat),
                           'dt': str(date)[:10],
                           'CostIN': float('{0:.2f}'.format(cost)),
                           'Nt': note,
                           # 'ProjectNm': pname,
                           'ApprovalStatus': 0}
                expense_entries.append(content)

        if save:
            json_to_file(expense_entries, 'entries.json')
        return expense_entries, float('{0:.2f}'.format(total_cost))

    #Validation to check whether submitted expenses exceed the per diem allotment
    def _per_diem_validation(self, daily_food_costs, cat, date, cost, note, per_diem, total_cost):
        
        # Try adding expenses to daily submission if the total is less than per diem amount
        # if the daily submission surpasses the per diem amount raise an error
        try:
            if date in daily_food_costs \
            and (daily_food_costs[date] + cost) <= int(per_diem) \
            and cat == 90:
                daily_food_costs[date] = daily_food_costs[date] + cost
                total_cost = total_cost + cost
            elif date in daily_food_costs \
            and (daily_food_costs[date] + cost) > int(per_diem) \
            and cat == 90:
                raise OverPerDiemError(date, per_diem, daily_food_costs, cost, note)
            elif (cost > int(per_diem)) \
            and cat == 90:
                raise OverPerDiemError(date, per_diem, daily_food_costs, cost, note)
            elif cat == 90:
                daily_food_costs[date] = cost
                total_cost = total_cost + cost

        # Handle raised OverPerDiemError by asking user for more information i.e. does the user want to continue?
        # How would they like to proceed? Options to proceed are to add remainder of budgeted per diem to 
        # submission and to add the complete item to the submission. If the user selects not to continue,
        # the latest expense is dropped completely.
        except OverPerDiemError:
            proceed = False
            while proceed is not True:
                proceed, total_cost, cost = self._complete_remainder_none(proceed, daily_food_costs, date, cost,\
                    note, per_diem, total_cost)

        return total_cost, cost, daily_food_costs 

    # Prompt user to specify how they'd like to handle the per diem overage (OverPerDiemError)
    def _complete_remainder_none(self, proceed, daily_food_costs, date, cost, note, per_diem, total_cost):
        
        # First prompt asks user to specify if they'd like to proceed with adding the current expense
        # to the submission
        proceed = input("Your expenses on {0} at {1} exceeded your per diem budget by {2:0.2f}. Do still you wish to proceed?"\
            "(y/n)\n\t".format(date, note, daily_food_costs[date] + cost - per_diem))
        
        if proceed is 'y':
            # Prompt user how to handle the over per diem expense. Option one is to submit an expense until the
            # per diem amount is reached. Option two is to add the complete expense to the report.
            complete_remainder = input("Would you prefer to use the remainder of your "\
                "budgeted per diem or to submit the complete expense (remainder/complete)\n\t")
            # If the user entered that they want to add the complete expense to the submission then
            # add the cost to the daily total and add the cost to the submission total and lastly alert
            # the user to their selection
            if complete_remainder == 'complete':
                daily_food_costs[date] = daily_food_costs[date] + cost
                total_cost = total_cost + cost
                print("You will be ${0:0.2f} over the allotted per diem!\n"\
                    .format(daily_food_costs[date] - per_diem))
                proceed = True
            # If the user entered that they want to add the remainder of expense until the per diem 
            # amount is reached to the submission then add the remainder to the daily total and add
            # the expense minus the remainder to the total cost                    
            elif complete_remainder == 'remainder':
                remainder = cost + daily_food_costs[date] - per_diem
                cost = cost - remainder             
                daily_food_costs[date] = daily_food_costs[date] + cost
                total_cost = total_cost + cost
                print("${0:0.2f} worth of expenses will be and added ${1:0.2f} of expenses won't be added!\n"\
                .format(cost, remainder))
                proceed = True
            else:
                print("Invalid entry, please enter 'remainder' or 'complete'\nStarting over...\n")
                proceed = False
        elif proceed is 'n':
            print("${} worth of expenses won't be added!\n".format(cost))
            proceed = True
        else:
            print("Invalid entry, please enter 'y' or 'n'\n")
            proceed = False
        return proceed, total_cost, cost

    def post_expenses(self, upload=False):
        expense_url = '{}/expense/detail'.format(BASE)
        expense_entries, total = self.prep_expenses()
        if upload is not True:
            input_map = {'Y': True, 'N': False}
            inp = input('Upload ${} {}{}\n\t'.format(total,
                                                   'worth of entries?',
                                                   '(y/n)')
                        ).upper()
            upload = input_map[inp]
        if upload:
            for entry in expense_entries:
                print(r.post(expense_url, headers=self.header,
                             data=json.dumps(entry).encode()),
                      entry['dt'], entry['CostIN'])
        else:
            print('${} expense entries not uploaded!'.format(total))
        return len(expense_entries)

    def get_active_reports(self):
        response = r.get('{0}/expense/reports'.format(BASE),
                         headers=self.header)
        print(response.status_code)
        return response.json()


def get_wb(workbook_name='Expenses.xlsx'):
    return load_workbook(filename=workbook_name, data_only=True)


def build_lookup_dictn_from_excel():
    """Build lookup_dictionaries from the excel workbook"""
    project_names = get_values('Projects', 'A2')
    project_ids = get_values('Projects', 'B2')
    category_names = get_values('Categories', 'A2')
    category_ids = get_values('Categories', 'B2')
    BT_LOOKUP['proj'] = dict(zip(project_ids, project_names))
    BT_LOOKUP['cat'] = dict(zip(category_ids, category_names))
    return project_ids, category_ids

def get_values(sheet_name, start, stop=None, workbook_name='Expenses.xlsx'):
    """Pulls a column (or section) of values from a Worksheet.
    Returns a list."""
    values = []
    sheet = get_wb(workbook_name)[sheet_name]
    if not stop:
        stop = sheet.max_row
    cells = [c[0].value for c in sheet[start:stop]]
    values = [c for c in cells if c is not None]
    return values


def get_picklist(auth_object, picklist_name):
    """Pulls a BigTime 'Picklist'
    Use to build project and expense catagory lookup tables.
    Requires Admin account."""
    # TODO: complete `get_picklist()` function
    valid_picklists = ['projects', 'ExpenseCodes']
    if picklist_name not in valid_picklists:
        raise ValueError('Not a valid picklist')
    pick_list_url = '{0}/picklist/{1}'.format(BASE, picklist_name)
    print(pick_list_url)
    response = r.get(pick_list_url, headers=auth_object.header)
    return response.json()
    # return response.json()


def json_to_file(json_obj, filename='data.json'):
    with open(filename, 'w') as f_out:
        json.dump(json_obj, f_out)
    return filename

class OverPerDiemError(Exception):

    def __init__(self, date, per_diem, daily_food_costs, cost, note):
        self.date = date
        self.per_diem = per_diem
        self.daily_food_costs = daily_food_costs
        self.cost = cost
        self.note = note

if __name__ == '__main__':
    print(__doc__)
    print('**DIR:', os.getcwd())
    print('*' * 79)
    
    still_running = True
    
    while still_running is True:
        excel = winclient.Dispatch("Excel.Application")
        excel.visible = True
        workbook = excel.Workbooks.open(os.getcwd() + '\Expenses.xlsx')
        up_to_date = input('Are the "Expenses" and "Setup" tabs up-to-date in the workbook? (y/n)').lower()
        if(up_to_date[0] == 'y'):
            workbook.Save()       
            workbook.Close()
            exp1 = Expensor()
            pp(exp1.post_expenses())
            still_running = False
    input('Process is complete, please press "Enter" to exit')
